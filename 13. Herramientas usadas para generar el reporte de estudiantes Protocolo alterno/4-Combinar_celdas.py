import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Cargar el archivo Excel en un DataFrame con pandas
df = pd.read_excel('./out/Reporte_estudiantes_convocatorias.xlsx')

# Guardar DataFrame en un nuevo archivo Excel temporal para poder cargarlo con openpyxl
df.to_excel('ruta_temporal.xlsx', index=False)

# Cargar el archivo temporal con openpyxl
wb = load_workbook('ruta_temporal.xlsx')
ws = wb.active

# Definir las columnas que deseas combinar
columnas_a_combinar = ['A', 'B', 'C', 'D', 'E']  # Cambia por las columnas que necesites (ejemplo: Nombre, Documento, Correo, etc.)

# Recorrer cada columna para combinar celdas con valores iguales
for col in columnas_a_combinar:
    col_idx = ws[col]  # Obtener las celdas de la columna
    start = 1  # Empezar desde la primera fila de datos
    prev_value = None  # Guardar el valor anterior

    for row in range(2, len(col_idx) + 1):  # Empezar desde la fila 2
        cell_value = ws[f"{col}{row}"].value

        # Comprobar si el valor actual es igual al valor anterior
        if cell_value == prev_value:
            # Si es igual, continuar
            continue
        else:
            # Si es diferente y hay más de una fila con el mismo valor anterior, combinar
            if row - start > 1:
                ws.merge_cells(f'{col}{start}:{col}{row-1}')
            
            # Actualizar el inicio y el valor anterior
            start = row
            prev_value = cell_value

    # Combinar las celdas restantes si el valor es igual hasta el final de la columna
    if len(col_idx) - start > 0:
        ws.merge_cells(f'{col}{start}:{col}{len(col_idx)}')

# Guardar los cambios en un nuevo archivo Excel
wb.save('./out/reporte_organizado.xlsx')
